import telebot
from telebot import types
import openpyxl
import os
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail




BOT_TOKEN ='6647530721:AAG4BFH3sf4ADFqzartPaWiwjwtu1B6boSo' #Bot token
SENDGRID_API_KEY = "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" #API key from SendGrid
bot = telebot.TeleBot(BOT_TOKEN)




#Start
@bot.message_handler(commands=['start'])     
def main(message):
    markup = types.ReplyKeyboardMarkup(row_width=1)
    buttons = [
        types.KeyboardButton("HR PartnerInterFreight"),
        types.KeyboardButton("Карьера в PartnerInterFreight"),
        types.KeyboardButton("Вакансии"),
        types.KeyboardButton("Наши Услуги"),
        types.KeyboardButton("Стажировка в PartnerInterFreight"),
        types.KeyboardButton("Отзывы сотрудников"),
        types.KeyboardButton("Наша команда"),
        types.KeyboardButton("Отправить резюме"),
        types.KeyboardButton("Контакты")
    ]
    markup.add(*buttons)
    
    hello_func = 'Вас приветствует HRPartnerInterFreightBOT! Ознакомиться с нашей Компанией вы можете, пройдя по ссылке http://partnerinterfreight.kz.tilda.ws/'
    bot.send_message(message.chat.id, hello_func, reply_markup=markup)
    
    

#Excel Table №1
def excel():    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(["ID", "Name", "Number", "Resume"])
    wb.save("Career.xlsx")
    
    
    
    
    
    
#Excel Table №2
def excel_internship():     
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Internship"
    ws.append(["ID", "Name", "Number"])
    wb.save("Internship.xlsx")
    
 
 
 
 
    
    
    
#Gmail-MESSAGE
def send_email(subject, to_email, name, number, resume):        
    try:
        mail_mess = f"Уважаемый HR PartnerInterFreight!Вы получили заявку от соискателя: {name} и номером: {number} в Telegram!Просим вас обработать заявку!"
        message = Mail(
            from_email='hrpifbot@gmail.com',
            to_emails=to_email,
            subject=subject,
            html_content=mail_mess
        )
        
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        
        return response.status_code
    except Exception as e:
        print("Ошибка при отправке электронной почты:", e)
        return None






#PhoneNumber    №1
def user_num(message, name):
    bot.send_message(message.chat.id, 'Контактный номер телефона:')
    bot.register_next_step_handler(message, lambda msg: user_resume(msg, name))
        




#PhoneNumber    №2
def user_internship(message):
    name = message.text  
    bot.send_message(message.chat.id, 'Введите ваш контактный номер телефона для стажировки:')
    bot.register_next_step_handler(message, final_internship_reg, name)




#Resume №1
def user_resume(message, name):
    resume = message.text
    bot.send_message(message.chat.id, 'Введите ссылку на ваше резюме:')
    bot.register_next_step_handler(message, final_reg, name, resume)





#Confirm Message    №1
def final_reg(message, name, resume):     
    try:
        number = message.text.strip()

        success = 'Вы были зарегистрированы'
        bot.send_message(message.chat.id, success)
        
        if not os.path.exists("Career.xlsx"):
            excel()
        
        wb = openpyxl.load_workbook("Career.xlsx")
        ws = wb.active
        ws.append([ws.max_row + 1, name, number, resume])
        wb.save("Career.xlsx")
        
        subject = "Регистрация на консультацию в телеграм"
        to_email = "anar.bairkhan@gmail.com"
        content = f"Уважаемый HR PartnerInterFreight! Вы получили заявку от СОИСКАТЕЛЯ: {name} и номером: {number} в Telegram!Просим вас обработать заявку!"

        try:
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            message = Mail(
                from_email='hrpifbot@gmail.com',
                to_emails=to_email,
                subject=subject,
                html_content=content
            )
            response = sg.send(message)
            status_code = response.status_code
            print("Статус отправки:", status_code)
        except Exception as e:
            print("Ошибка при отправке электронной почты:", e)
        
    except Exception as e:
        error_message = f'Произошла ошибка: {e}. Пожалуйста, попробуйте позже.'
        bot.send_message(message.chat.id, error_message)





#Confirm Message    №2
def final_internship_reg(message, name):       
    try:
        number = message.text.strip()

        success = 'Вы были зарегистрированы на стажировку'
        bot.send_message(message.chat.id, success)

        if not os.path.exists("Internship.xlsx"):
            excel_internship()

        wb = openpyxl.load_workbook("Internship.xlsx")
        ws = wb.active
        ws.append([ws.max_row + 1, name, number])
        wb.save("Internship.xlsx")

        subject = "Регистрация на стажировку в телеграм"
        to_email = "anar.bairkhan@gmail.com"
        content = f"Уважаемый HR PartnerInterFreight! Вы получили заявку на СТАЖИРОВКУ от: {name} и номером: {number} в Telegram! Просим вас обработать заявку."

        try:
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            message = Mail(
                from_email='hrpifbot@gmail.com',
                to_emails=to_email,
                subject=subject,
                html_content=content
            )
            response = sg.send(message)
            status_code = response.status_code
            print("Статус отправки:", status_code)
        except Exception as e:
            print("Ошибка при отправке электронной почты:", e)

    except Exception as e:
        error_message = f'Произошла ошибка: {e}. Пожалуйста, попробуйте позже.'
        bot.send_message(message.chat.id, error_message)





#Name   №1
@bot.message_handler(func=lambda message: message.text == "Карьера в PartnerInterFreight")
def user_name(message):
    career_text = (
        "Мы активно развиваем новые направления деятельности, поэтому у каждого сотрудника есть возможности для применения и развития своих профессиональных качеств и построения карьеры.\n"
        "ПРИСОЕДИНЯЙТЕСЬ К ТРАНСПОРТНО-ЛОГИСТИЧЕСКОЙ КОМПАНИИ\n"
        "Развивайтесь в карьере и получайте премии!\n"
        "Должностной оклад от 400 000 тг + бонусы, подарки на день рождения\n"
        "Лучшее внутреннее обучение и карьерный рост\n"
        "Комфортный офис в центре Алматы и участие в 5 корпоративных мероприятий в год.\n\n"
        "Если ты хочешь быть частью нашей команды, оставь заявку.\n\n"
        'Укажите ФИО:'
    )
    bot.send_message(message.chat.id, career_text)
    bot.register_next_step_handler(message, lambda msg: user_num(msg, message.text))



#Name   №2
@bot.message_handler(func=lambda message: message.text == "Стажировка в PartnerInterFreight")
def internship_handler(message):
    internship_text = (
        "Если ты выпускник ВУЗа по направлению транспортной логистики, "
        "то ты можешь пройти Стажировку в нашей Компании.\n"
        "Для этого оставь заявку.\n\n"
        "Укажите ФИО:"
    )
    bot.send_message(message.chat.id, internship_text)
    bot.register_next_step_handler(message, user_internship)



#Phot0
@bot.message_handler(func=lambda message: message.text == "Наша команда")
def pif_team(message):
    photo = open('pifpic.jpeg', 'rb')
    caption = (
        "В нашей Команде сотрудники, работающие с удовольствием и чувствующие единение с Компанией. Наши сотрудники – это «амбассадоры» нашей Компании!\n"
        "Корпоративная культура Компании основана на взаимном уважении и доверии.\n"
        "Каждый день в рабочей и дружественной атмосфере МЫ:\n"
        "• Ориентированы на результат;\n"
        "• Развиваемся и обучаемся;\n"
        "• Клиентоориентированы с коллегами и клиентами;\n"
        "• Получаем высокое вознаграждение за лучшие результаты.\n"
        "Если ты созвучен с нашими ценностями и принципами, готов достигать профессиональных вершин, то у тебя есть шанс быть в НАШЕЙ КОМАНДЕ!"
    )
    bot.send_photo(message.chat.id, photo, caption=caption)



#Review
@bot.message_handler(func=lambda message: message.text == "Отзывы сотрудников")
def review(message):
    text = (
        "Наши сотрудники делятся своими отзывами:\n\n"
        "<b>Анищенко Анастасия, Бухгалтер-операционист:</b>\n"
        "«Работа в компании дала мне возможность развиваться профессионально, улучшить навыки и получить ценный опыт.\n"
        "Я очень благодарна компании за предоставленные возможности для развития и роста, за их поддержку, мудрые советы и вдохновение. Они создали стимулирующую рабочую среду, где каждый сотрудник ценится и мотивируется достигать своих целей. Здесь ценится инициатива сотрудников, каждый вносит свой вклад в развитие компании.»\n\n"
        "<b>Шарафутдинова Зарина, Руководитель отдела развития бизнеса:</b>\n"
        "«Я работаю здесь уже в течение трех месяцев, и не могу не подчеркнуть, как положительно это отразилось на качестве моей жизни. Каждый рабочий день у нас проходит интересно и увлекательно благодаря динамичной атмосфере в коллективе.\n"
        "Одним из основных преимуществ работы в нашей компании является современный подход к бизнесу. В компании также предлагаются отличные возможности для карьерного роста. Мы поддерживаем и поощряем амбиции наших сотрудников.\n"
        "Если вы стремитесь к росту и развитию, любите делиться идеями и видеть результаты своей работы, у нас вы найдете множество возможностей для самореализации.»\n\n"
        "<b>Кульдурбаев Бахытжан, Ведущий менеджер по работе с поставщиками:</b>\n"
        "«Начиная работать в компании «PartnerInterFreight», я ощутил свою значимость, что не сказал бы о предыдущей моей работе. На мою жизнь повлияло тем, что начал много времени уделять семье и личному саморазвитию.\n"
        "Нравится то, что компания даёт тебе возможность самореализовываться, как специалисту. Чувствуешь себя, как дома, как в кругу семьи. Все плюсы конечно не перечислишь, но самый очевидный «+» это то, что всё «Честно и Прозрачно», по сравнению с компаниями где я работал.\n"
        "В компании есть карьерный и личностный рост. Я пришел на вакансию менеджера отдела, а по истечению и результату испытательного срока, меня перевели на должность Ведущего менеджера отдела по работе с поставщиками»"
    )
    bot.send_message(message.chat.id, text, parse_mode='HTML')



#Contacts-Links
@bot.message_handler(func=lambda message: message.text == "Контакты")
def cont(message):
    links_info = (
        "+7 (701) 327 32 97 HR менеджер\n"
        "+7 (727) 346 87 99 Офис-менеджер\n"
        "info@piftec.kz\n"
        "Адрес офиса:\n"
        "050042, РК, г. Алматы, ул. Жандосова 98, 7 этаж, офис 700\n\n"
        "Следите за нами:\n"
        "🎥 YouTube: https://www.youtube.com/@PartnerInterFreightCompany\n"
        "📸 Instagram: https://www.instagram.com/partnerinterfreight/?igshid=MzRlODBiNWFlZA%3D%3D"
    )
    
    inline_markup = types.InlineKeyboardMarkup()
    youtube_button = types.InlineKeyboardButton("🎥     YouTube", url="https://www.youtube.com/@PartnerInterFreightCompany")
    instagram_button = types.InlineKeyboardButton("📸 Instagram", url="https://www.instagram.com/partnerinterfreight/?igshid=MzRlODBiNWFlZA%3D%3D")
    pif_button = types.InlineKeyboardButton("🌐 Website", url="info@piftec.kz")
    inline_markup.add(youtube_button, instagram_button, pif_button)
    
    message_text = f"Контакты и соцсети:\n{links_info}"
    
    bot.send_message(message.chat.id, message_text, reply_markup=inline_markup)



#Buttons
@bot.message_handler(func=lambda message: True)    
def click(message):
    try:
        response = {
            'Вакансии': 'Ознакомиться с вакансиями Компании вы можете здесь: https://hh.kz/search/vacancy?from=employerPage&employer_id=2339821',
            'Наши Услуги': "1.	ЖЕЛЕЗНОДОРОЖНЫЕ ПЕРЕВОЗКИ\n"
"•	Подбор и предоставление подвижного состава;\n"
"•	Расчёт и оплата железнодорожных тарифов;\n"
"•	Оказание услуг грузополучателя/грузоотправителя;\n"
"•	Отслеживание движения груза.\n\n"

"2.	МУЛЬТИМОДАЛЬНЫЕ ПЕРЕВОЗКИ\n"
"•	Разработка маршрутов поставки;\n"
"•	Расчёт ставок и оптимизация транспортных издержек;\n"
"•	Организация перевозок всеми видами транспорта;\n"
"•	Организация терминальных услуг;\n"
"•	Отслеживание движения груза;\n"
"•	Оформление транспортных документов.\n\n"
"3.	АВИАПЕРЕВОЗКИ\n"
"•	Забор груза непосредственно со склада отправителя;\n"
"•	Авиаперевозка генеральных, скоропортящихся, опасных, негабаритных и тяжеловесных грузов;\n"
"•	Оформление транспортных документов;\n"
"•	Срочная доставка грузов (1-3 дня);\n"
"•	Организация грузовых чартерных рейсов.\n\n"
"4.	АВТОМОБИЛЬНЫЕ ПЕРЕВОЗКИ\n"
"•	Организация автомобильных перевозок в международном направлении отдельными машинами и в составе консолидации;\n"
"•	Оформление транспортных документов;\n"
"•	Отслеживание движения груза;\n"
"•	Терминальная обработка.",
            "Галлерея": "Команда PIF",
            "Отзывы сотрудников": "",
            'Отправить резюме': 'Необходимо отправить ваше резюме на электронную почту по адресу: g.abyken@piftec.kz',
            "Карьера в PartnerInterFreight": (
                "Мы активно развиваем новые направления деятельности, поэтому у каждого сотрудника есть возможности для применения и развития своих профессиональных качеств и построения карьеры.\n"
                "ПРИСОЕДИНЯЙТЕСЬ К ТРАНСПОРТНО-ЛОГИСТИЧЕСКОЙ КОМПАНИИ. Развивайтесь в карьере и получайте премии!\n"
                "Должностной оклад от 400 000 тг + бонусы, подарки на день рождения\n"
                "Лучшее внутреннее обучение и карьерный рост\n"
                "Комфортный офис в центре Алматы и участие в 5 корпоративных мероприятий в год.\n\n"
                "Если ты хочешь быть частью нашей команды, оставь заявку:"
            ),
            "HR PartnerInterFreight": (
                "HR PartnerInterFreight предлагает нашим сотрудникам ЛУЧШЕЕ ВНУТРЕННЕЕ ОБУЧЕНИЕ!\n\n"
                "Мы бесплатно обучим вас всей системе бизнеса, укрепим личностные навыки и превратим в высококлассного специалиста.\n\n"
                "В нашей Компании каждый сотрудник знает что, как и для чего ему нужно делать. Вам не придется сомневаться – глубокое обучение, должностные инструкции и прозрачная система компании максимально упрощают работу.\n\n"
                "Связаться с HR PartnerInterFreight можно по телефону +7 701 327 3297 или по электронной почте g.abyken@piftec.kz"
            ),
            "Стажировка в PartnerInterFreight": (
                "Если ты выпускник ВУЗа по направлению транспортной логистики, то ты можешь пройти Стажировку в нашей Компании.\n"
                "Для этого оставь заявку:"
            )
        }
        
        bot.send_message(message.chat.id, response.get(message.text, "Выберите какой вопрос вас интересует в разделе меню:"))
    except Exception as e:
        bot.send_message(message.chat.id, "Произошла ошибка при обработке вашего ввода. Пожалуйста, воспользуйтесь доступными кнопками или задайте вопрос более конкретно.")
        print("Ошибка при обработке ввода:", e)    
    


#Always Code
if __name__ == "__main__":
    try:
        excel()
        excel_internship()
        bot.polling(none_stop=True)
    except Exception as e:
        print("Ошибка:", e)